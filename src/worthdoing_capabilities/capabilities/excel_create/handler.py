"""Excel document creation capability handler."""
from pathlib import Path


async def execute(input_data: dict) -> dict:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output_path = Path(input_data["output_path"])
    sheets_data = input_data["sheets"]

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_rows = 0

    for sheet_def in sheets_data:
        ws = wb.create_sheet(title=sheet_def.get("name", "Sheet"))

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        headers = sheet_def.get("headers", [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        rows = sheet_def.get("rows", [])
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        total_rows += len(rows)

        # Set column widths
        column_widths = sheet_def.get("column_widths", [])
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Auto-fit if no widths specified
        if not column_widths:
            for col_idx in range(1, len(headers) + 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 2)),
                    default=10,
                )
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if headers:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return {
        "path": str(output_path),
        "sheets_count": len(sheets_data),
        "total_rows": total_rows,
    }
